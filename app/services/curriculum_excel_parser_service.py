from openpyxl import load_workbook
from app.db.mongo import curriculum_collection
import time


class CurriculumExcelParserService:

    @staticmethod
    def parse_and_save(file_stream):
        wb = load_workbook(file_stream, data_only=True)

        # -----------------------------
        # Read Metadata Sheet
        # -----------------------------
        meta = wb["Curriculum Info"]

        class_name = meta["B1"].value
        department = meta["B2"].value
        total_semesters = int(meta["B3"].value)
        effective_year = meta["B4"].value

        semester_data = {}
        grand_total_credits = 0

        # -----------------------------
        # Parse Each Semester Sheet
        # -----------------------------
        for sheet_name in wb.sheetnames:
            if sheet_name == "Curriculum Info":
                continue

            sheet = wb[sheet_name]
            semester_courses = []
            semester_credits = 0

            # Data starts from row 4
            row = 4
            while True:
                code = sheet[f"A{row}"].value
                name = sheet[f"B{row}"].value

                if not code or not name:
                    break

                l = sheet[f"C{row}"].value or 0
                t = sheet[f"D{row}"].value or 0
                p = sheet[f"E{row}"].value or 0

                category = code[:2]

                # Credit Calculation
                if category in ["MC", "AU"]:
                    credits = 0
                else:
                    credits = (1 * l) + (1 * t) + (0.5 * p)

                semester_credits += credits
                grand_total_credits += credits

                semester_courses.append({
                    "name": name,
                    "code": code,
                    "category": category,
                    "l": l,
                    "t": t,
                    "p": p,
                    "credits": credits
                })

                row += 1

            semester_data[sheet_name] = {
                "courses": semester_courses,
                "semesterCredits": semester_credits
            }

        # -----------------------------
        # Final Curriculum Document
        # -----------------------------
        curriculum_doc = {
            "className": class_name,
            "department": department,
            "totalSemesters": total_semesters,
            "effectiveYear": effective_year,
            "semesterData": {
                k: v["courses"] for k, v in semester_data.items()
            },
            "totalCredits": grand_total_credits,
            "uploadedAt": int(time.time())
        }

        curriculum_collection.insert_one(curriculum_doc)

        return {
            "className": class_name,
            "department": department,
            "totalCredits": grand_total_credits,
            "semesterWiseCredits": {
                k: v["semesterCredits"] for k, v in semester_data.items()
            }
        }
