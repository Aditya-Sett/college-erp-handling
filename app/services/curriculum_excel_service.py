from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
from app.utils.semester_naming import semester_name

class CurriculumExcelService:
    @staticmethod
    def generate_excel(classname, department, total_semesters, effective_year):
        wb = Workbook()

        # -------------------------
        # Sheet 1: Metadata Sheet
        # -------------------------
        meta_sheet = wb.active
        meta_sheet.title = "Curriculum Info"

        meta_sheet["A1"] = "Class Name"
        meta_sheet["B1"] = classname

        meta_sheet["A2"] = "Department"
        meta_sheet["B2"] = department

        meta_sheet["A3"] = "Total Semesters"
        meta_sheet["B3"] = total_semesters

        meta_sheet["A4"] = "Effective Year"
        meta_sheet["B4"] = effective_year

        # Make headers bold
        for row in range(1, 5):
            meta_sheet[f"A{row}"].font = Font(bold=True)

        # Protect sheet (read-only)
        meta_sheet.protection.sheet = True

        # -------------------------
        # Semester Sheets
        # -------------------------
        for sem in range(1, total_semesters + 1):
            sheet = wb.create_sheet(title=semester_name(sem))

            # ---- Merge Title Row ----
            sheet.merge_cells("A1:E1")
            sheet["A1"] = semester_name(sem)
            sheet["A1"].font = Font(bold=True)
            sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

            # ---- Header Structure ----
            sheet.merge_cells("A2:A3")
            sheet["A2"] = "Course Code"

            sheet.merge_cells("B2:B3")
            sheet["B2"] = "Course Name"

            sheet.merge_cells("C2:E2")
            sheet["C2"] = "Contact Hours"

            sheet["C3"] = "L"
            sheet["D3"] = "T"
            sheet["E3"] = "P"

            # ---- Styling Headers ----
            header_cells = ["A2", "B2", "C2", "C3", "D3", "E3"]
            for cell in header_cells:
                sheet[cell].font = Font(bold=True)
                sheet[cell].alignment = Alignment(horizontal="center", vertical="center")

            # ---- Freeze Top 3 Rows ----
            sheet.freeze_panes = "A4"

        return wb